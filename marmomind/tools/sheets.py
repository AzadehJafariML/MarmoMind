# MarmoMind: AI Agent designed by Azadeh Jafari (jfr.azadeh@gmail.com) for the
# Everling Lab, Centre for Functional and Metabolic Mapping, University of
# Western Ontario. Created May 2026.
"""Sheet tools (v1 backend = local xlsx via openpyxl).

The same signatures swap to a live Google Sheet (service account) later with no
change to the agent. Set env MARMOMIND_XLSX to point at a copy for testing.
"""
import re
from pathlib import Path
from typing import Any

import openpyxl
from claude_agent_sdk import tool

from ..common import ok, fail, load_settings, data_path
from ..matching import resolve_monkey

# Per-monkey tab columns, in order.
COLUMNS = ["Session", "Date of scan", "State of the animal", "Time In", "Time Out",
          "Run", "Volumes", "Resolution", "TR", "Coils", "Task", "Comments"]


def _xlsx_path() -> Path:
    # data_path("xlsx") already honors the MARMOMIND_XLSX env override.
    return data_path("xlsx")


def lookup_id(subject: str) -> dict:
    """Resolve a subject to its monkey ID + per-monkey tab, TOLERANT of formatting
    but honest about ambiguity (never guesses between two animals)."""
    path = _xlsx_path()
    wb = openpyxl.load_workbook(path)
    summary = wb[load_settings()["sheet"]["summary_tab"]]
    header = [c.value for c in summary[1]]
    try:
        a_i, id_i = header.index("Animal name"), header.index("ID")
    except ValueError:
        return {"found": False, "subject": subject, "error": "Summary header missing Animal name/ID"}
    sn_i = header.index("Sheet name") if "Sheet name" in header else None

    candidates = [
        {"name": row[a_i], "id": (str(row[id_i]).strip() if row[id_i] is not None else None),
         "sheet_name": (row[sn_i] if sn_i is not None else None)}
        for row in summary.iter_rows(min_row=2, values_only=True) if row[a_i]
    ]
    res = resolve_monkey(subject, candidates)
    if res["status"] != "matched":
        # ambiguous / unmatched -> surface the flag; found=False so the agent defers
        return {"found": False, "subject": subject, **res}

    mid, name = res["id"], str(res["name"]).strip()
    tab = f"{mid}_{name}"
    if tab not in wb.sheetnames and res.get("sheet_name") \
            and str(res["sheet_name"]).strip() in wb.sheetnames:
        tab = str(res["sheet_name"]).strip()
    return {"found": True, "status": "matched", "subject": subject, "id": mid,
            "matched_name": name, "tab": tab, "tab_exists": tab in wb.sheetnames}


def _session_num(v):
    """Parse a Session cell to an int. The sheet uses BOTH 's105' strings and
    bare 105 integers across its history; accept either. Non-session -> None."""
    if v in (None, ""):
        return None
    m = re.match(r"^s?(\d+)$", str(v).strip(), re.IGNORECASE)
    return int(m.group(1)) if m else None


def next_session(tab: str) -> dict:
    wb = openpyxl.load_workbook(_xlsx_path())
    if tab not in wb.sheetnames:
        return {"error": f"tab '{tab}' not found"}
    ws = wb[tab]
    last = None          # LAST non-blank session value going top-down
    last_raw = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        n = _session_num(row[0])
        if n is not None:
            last, last_raw = n, row[0]
    nxt = "s1" if last is None else f"s{last + 1}"
    return {"next_session": nxt, "last_seen": (f"s{last}" if last is not None else None),
            "last_raw": last_raw}


def append_rows(tab: str, session: str, rows: list) -> dict:
    path = _xlsx_path()
    wb = openpyxl.load_workbook(path)
    if tab not in wb.sheetnames:
        return {"error": f"tab '{tab}' not found"}
    ws = wb[tab]
    header = [(c.value or "") for c in ws[1]]

    def col_index(name: str) -> int:
        return header.index(name) if name in header else COLUMNS.index(name)

    # ws.max_row overcounts when a tab declares trailing blank rows, which would
    # leave a gap. Find the TRUE last row that holds any data and append after it.
    last_data = 1
    for r in range(ws.max_row, 0, -1):
        if any(c.value not in (None, "") for c in ws[r]):
            last_data = r
            break
    start = last_data + 1
    written = []
    for i, r in enumerate(rows):
        rownum = start + i
        rec = dict(r)
        rec["Session"] = session if i == 0 else ""   # session on first run-row only
        for name, val in rec.items():
            if name in COLUMNS or name in header:
                ws.cell(row=rownum, column=col_index(name) + 1,
                        value=("" if val is None else val))
        written.append(rownum)
    wb.save(path)
    return {"appended": len(rows), "rows": written, "tab": tab, "session": session,
            "xlsx": str(path)}


@tool(
    "lookup_monkey_id",
    "Open the Summary tab and match an Animal name to its ID and per-monkey tab "
    "name (e.g. 'After' -> id 'm16', tab 'm16_After'). Read-only.",
    {"subject": str},
)
async def lookup_monkey_id(args: dict[str, Any]) -> dict[str, Any]:
    return ok(lookup_id(args["subject"]))


@tool(
    "get_next_session_number",
    "Read the per-monkey tab's Session column, find the LAST non-blank 's<N>' "
    "value and increment. The agent is NOT the only writer (anatomical-only visits "
    "are entered by humans), so NEVER renumber from a processed-run count — only "
    "continue the column's existing sequence. Read-only.",
    {"tab": str},
)
async def get_next_session_number(args: dict[str, Any]) -> dict[str, Any]:
    return ok(next_session(args["tab"]))


@tool(
    "append_run_rows",
    "Append one row PER FUNCTIONAL run to the per-monkey tab (never the ap). "
    "Session label 's<N>' on the FIRST run-row only; blank on continuation rows. "
    "Missing value -> blank cell, never stop. Each row is a dict keyed by the tab "
    "columns (omit 'Session' — the tool places it). MUTATING — requires approval.",
    {"tab": str, "session": str, "rows": list},
)
async def append_run_rows(args: dict[str, Any]) -> dict[str, Any]:
    res = append_rows(args["tab"], args["session"], args["rows"])
    return fail(res["error"], **res) if "error" in res else ok(res)
